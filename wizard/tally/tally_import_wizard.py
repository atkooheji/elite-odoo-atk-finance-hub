import xml.etree.ElementTree as ET
import base64
import io
import logging
import re
import openpyxl
import pandas as pd
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

class TallyImportWizard(models.TransientModel):
    _name = 'tally.import.wizard'
    _description = 'Tally Import Wizard'

    import_file = fields.Binary(string='Tally File', required=True)
    file_name = fields.Char(string='File Name')
    file_type = fields.Selection([('excel', 'Excel (DayBook / TB)'), ('xml', 'XML (Masters / Vouchers)')], default='excel', required=True)
    import_type = fields.Selection([
        ('masters', 'Masters'),
        ('daybook', 'Day Book'),
        ('trialbalance', 'Trial Balance'),
        ('opening', 'Opening Balances'),
        ('loan', 'Loan Schedule'),
        ('ageing', 'Ageing')
    ], default='daybook', required=True)
    
    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date')
    
    journal_id = fields.Many2one('account.journal', string='Target Journal', domain=[('type', 'in', ['general', 'bank', 'cash', 'credit', 'credit_card'])])
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    balancing_account_id = fields.Many2one('account.account', string='Balancing Account', help="Account to use if the trial balance is not balanced.")
    auto_post = fields.Boolean(string='Post Directly', default=False)
    skip_unbalanced = fields.Boolean(string='Import Unbalanced as Draft', default=True)
    auto_balance = fields.Boolean(string='Auto-Balance Entry', default=True)
    is_opening_entry = fields.Boolean(string='Is Opening Entry', default=False, help="If checked, prioritized Opening Balance columns from Tally.")
    only_closing_balance = fields.Boolean(string='Closing Balance Only (TB)', default=False)
    follow_grouping = fields.Boolean(string='Follow Tally Grouping', default=True)
    import_tally_groups = fields.Boolean(string='Import Tally Groups', default=False)
    skip_group_totals = fields.Boolean(string='Skip Group Totals', default=True)
    one_move_per_group = fields.Boolean(string='One Entry Per Group', default=False)
    batch_id = fields.Many2one('tally.import.batch', string='Current Batch', readonly=True)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('analysis', 'Analysis'),
        ('done', 'Done')
    ], default='draft', required=True)

    def action_import(self):
        self.ensure_one()
        if not self.import_file: raise UserError(_("Please upload a file."))
        file_content = base64.b64decode(self.import_file)
        if self.file_type == 'xml': return self._process_xml(file_content)
        return self._process_excel(file_content)

    def action_delete_data(self):
        """ Restoration: Wipe existing Tally data in the selected date range. """
        if not self.date_from or not self.date_to:
            raise UserError(_("Please select a date range to wipe data."))
        # Safety: Only delete moves with 'is_tally_import'
        moves = self.env['account.move'].search([
            ('date', '>=', self.date_from),
            ('date', '<=', self.date_to),
            ('is_tally_import', '=', True)
        ])
        if moves:
            moves.button_draft()
            moves.unlink()
        return True

    def _process_excel(self, file_content):
        try: df = pd.read_excel(io.BytesIO(file_content), header=None, dtype=str)
        except Exception as e: raise UserError(_("Error reading Excel file: %s") % str(e))
        if self.import_type == 'daybook': return self._process_daybook(df)
        elif self.import_type in ['trialbalance', 'opening']: return self._process_trial_balance(file_content)
        elif self.import_type == 'loan': return self._process_loan_schedule(df)
        elif self.import_type == 'ageing': return self._process_ageing_report(df)
        raise UserError(_("Excel import type not supported yet."))

    def _parse_amount(self, val):
        if pd.isna(val) or not val: return 0.0
        num_part = re.sub(r'[^0-9\.\-]', '', str(val).lower().replace(',', ''))
        try: return float(num_part)
        except: return 0.0

    def _process_daybook(self, df):
        header_row = -1
        # Use fuzzy search for headers
        for i, row in df.iterrows():
            row_vals = [str(x).strip().lower() for x in row.tolist() if pd.notnull(x)]
            has_part = any('particular' in v for v in row_vals)
            has_amt = any('debit' in v or 'credit' in v or 'amount' in v for v in row_vals)
            if has_part and has_amt:
                header_row = i
                break
        if header_row == -1: raise UserError(_("Invalid DayBook header. Could not find 'Particulars' or 'Amount' columns."))

        # Identify actual column names from found header row
        header_vals = [str(c).strip().lower() if pd.notnull(c) else "nan" for c in df.iloc[header_row]]
        df.columns = header_vals
        data_df = df.iloc[header_row+1:].copy()
        
        vch_no_col = next((c for c in df.columns if 'vch no' in c or 'voucher no' in c), 'nan')
        date_col = next((c for c in df.columns if 'date' in c), 'nan')
        vch_type_col = next((c for c in df.columns if 'vch type' in c or 'voucher type' in c), 'nan')
        part_col = next((c for c in df.columns if 'particular' in c), 'nan')
        debit_col = next((c for c in df.columns if 'debit' in c), 'nan')
        credit_col = next((c for c in df.columns if 'credit' in c), 'nan')

        data_df[vch_no_col] = data_df[vch_no_col].ffill()
        data_df[date_col] = data_df[date_col].ffill()
        if vch_type_col in data_df.columns: data_df[vch_type_col] = data_df[vch_type_col].ffill()

        # 1. Handle Account Mappings
        data_df[part_col] = data_df[part_col].fillna('')
        all_file_ledgers = [str(l).strip() for l in data_df[part_col].unique().tolist() if str(l).strip() and str(l).strip().lower() not in ['total', 'grand total', 'nan', '']]

        mapping_model = self.env['tally.account.mapping']
        acc_model = self.env['account.account']
        for ledger in all_file_ledgers:
            if not mapping_model.search([('tally_ledger_name', '=', ledger), ('company_id', '=', self.company_id.id)], limit=1):
                acc_fields = acc_model._fields
                acc_domain = [('name', '=', ledger)]
                if 'company_id' in acc_fields: acc_domain.append(('company_id', '=', self.company_id.id))
                elif 'company_ids' in acc_fields: acc_domain.append(('company_ids', 'in', [self.company_id.id]))
                
                odoo_acc = acc_model.search(acc_domain, limit=1)
                mapping_model.create({'tally_ledger_name': ledger, 'odoo_account_id': odoo_acc.id if odoo_acc else False, 'company_id': self.company_id.id})

        unmapped_accs = mapping_model.search([('tally_ledger_name', 'in', all_file_ledgers), ('odoo_account_id', '=', False), ('company_id', '=', self.company_id.id)])
        
        # 2. Handle Voucher Type Mappings
        all_file_vchs = []
        if vch_type_col in data_df.columns:
            all_file_vchs = [str(v).strip() for v in data_df[vch_type_col].unique().tolist() if str(v).strip() and str(v).strip().lower() != 'nan']
        
        vch_mapping_model = self.env['tally.voucher.mapping']
        journal_model = self.env['account.journal']
        for vch in all_file_vchs:
            if not vch_mapping_model.search([('tally_voucher_type', '=', vch), ('company_id', '=', self.company_id.id)], limit=1):
                odoo_journal = journal_model.search([('name', '=', vch), ('company_id', '=', self.company_id.id)], limit=1)
                if not odoo_journal:
                    odoo_journal = journal_model.search([('type', '=', 'general'), ('company_id', '=', self.company_id.id)], limit=1)
                vch_mapping_model.create({'tally_voucher_type': vch, 'odoo_journal_id': odoo_journal.id if odoo_journal else False, 'company_id': self.company_id.id})
        
        unmapped_vchs = vch_mapping_model.search([('tally_voucher_type', 'in', all_file_vchs), ('odoo_journal_id', '=', False), ('company_id', '=', self.company_id.id)])

        # Redirect if missing mappings
        if unmapped_accs or unmapped_vchs:
            if unmapped_accs:
                return {'name': _('Complete Account Mappings'), 'type': 'ir.actions.act_window', 'res_model': 'tally.account.mapping', 'view_mode': 'list', 'domain': [('id', 'in', unmapped_accs.ids)], 'target': 'current'}
            else:
                return {'name': _('Complete Voucher Mappings'), 'type': 'ir.actions.act_window', 'res_model': 'tally.voucher.mapping', 'view_mode': 'list', 'domain': [('id', 'in', unmapped_vchs.ids)], 'target': 'current'}

        mappings = mapping_model.search([('tally_ledger_name', 'in', all_file_ledgers), ('company_id', '=', self.company_id.id)])
        ledger_map = {m.tally_ledger_name: m.odoo_account_id.id for m in mappings}
        
        vch_mappings = vch_mapping_model.search([('tally_voucher_type', 'in', all_file_vchs), ('company_id', '=', self.company_id.id)])
        vch_map = {m.tally_voucher_type: m.odoo_journal_id.id for m in vch_mappings}

        batch = self.batch_id or self.env['tally.import.batch'].create({'file_name': self.file_name or 'Tally Excel', 'company_id': self.company_id.id})
        created_moves = self.env['account.move']
        
        grouped = data_df.groupby(vch_no_col)
        for vch_no, group in grouped:
            if not vch_no or str(vch_no).lower() == 'nan': continue
            line_ids = []
            for idx, row in group.iterrows():
                particulars = str(row.get(part_col, '')).strip()
                if particulars not in ledger_map: continue
                debit, credit = self._parse_amount(row.get(debit_col)), self._parse_amount(row.get(credit_col))
                if debit == 0 and credit == 0: continue
                line_ids.append((0, 0, {'account_id': ledger_map[particulars], 'name': particulars, 'debit': debit, 'credit': credit}))
            
            if not line_ids: continue
            vch_type = str(group.iloc[0].get(vch_type_col, '')).strip()
            target_journal = vch_map.get(vch_type) or self.journal_id.id or self.env['account.journal'].search([('type', '=', 'general')], limit=1).id
            
            move_context = {'check_move_validity': False} if self.skip_unbalanced else {}
            move = self.env['account.move'].with_context(**move_context).create({
                'move_type': 'entry', 'journal_id': target_journal,
                'date': group.iloc[0].get(date_col) or fields.Date.today(), 'ref': vch_no, 'line_ids': line_ids, 'tally_batch_id': batch.id, 'is_tally_import': True
            })
            created_moves |= move
            if self.auto_post and not self.skip_unbalanced:
                try: move.action_post()
                except: pass

        batch.write({'state': 'done'})
        return {'name': _('Import Successful'), 'type': 'ir.actions.act_window', 'res_model': 'tally.import.batch', 'view_mode': 'form', 'res_id': batch.id, 'target': 'current'}

    def _parse_tally_amount(self, val):
        if pd.isna(val) or not val: return 0.0, None
        s = str(val).lower().strip()
        side = None
        if 'dr' in s: side = 'dr'
        elif 'cr' in s: side = 'cr'
        num_part = re.sub(r'[^0-9\.\-]', '', s.replace(',', ''))
        try: return float(num_part), side
        except: return 0.0, None

    def _process_trial_balance(self, file_content):
        import openpyxl
        import io
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_("Could not read the Excel file: %s") % str(e))
            
        header_row = -1
        col_map = {}
        target_import = 'opening' if (self.import_type == 'opening' or self.is_opening_entry) else 'closing'

        # 1. Find Header Row and Columns
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            row_vals = [str(cell.value).strip().lower() if cell.value else '' for cell in row]
            if any('particular' in v for v in row_vals):
                header_row = row_idx
                for r_offset in range(3): 
                    if header_row + r_offset > ws.max_row: break
                    for col_idx, cell in enumerate(ws[header_row + r_offset]):
                        val = str(cell.value).strip().lower() if cell.value else ''
                        if not val: continue
                        if 'particular' in val and 'particulars' not in col_map: col_map['particulars'] = col_idx
                        elif target_import == 'opening' and 'opening' in val and 'balance' in val: col_map['amount_col'] = col_idx
                        elif target_import == 'closing' and ('closing' in val or 'balance' in val) and 'amount_col' not in col_map: col_map['amount_col'] = col_idx
                        elif 'debit' in val and 'debit' not in col_map: col_map['debit'] = col_idx
                        elif 'credit' in val and 'credit' not in col_map: col_map['credit'] = col_idx
                break
                
        if header_row == -1: raise UserError(_("Invalid TB header. 'Particulars' column not found."))
        
        has_amt_col = 'amount_col' in col_map
        has_dr_cr = 'debit' in col_map and 'credit' in col_map
        
        if not has_amt_col and not has_dr_cr:
            raise UserError(_("Could not identify the Balance columns. Expected 'Opening Balance', 'Closing Balance', or 'Debit/Credit'."))

        valid_rows = []
        all_accounts = []
        part_idx = col_map.get('particulars', 0)
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=part_idx+1)
            name = str(name_cell.value).strip() if name_cell.value else ''
            
            # Skip empty or header-like names
            if not name or name.lower() in ['particulars', 'none']: continue
            
            # Implementation of Skip Summary/Total Lines
            is_total_keyword = any(k in name.lower() for k in ['total', 'grand total', 'subtotal'])
            is_bold = name_cell.font and name_cell.font.bold
            
            if self.skip_group_totals:
                if is_total_keyword or is_bold:
                    continue
            
            debit_amt, credit_amt = 0.0, 0.0
            if has_amt_col:
                cell = ws.cell(row=row_idx, column=col_map['amount_col']+1)
                amt, side = self._parse_tally_amount(cell.value)
                
                if amt == 0: continue
                
                # Smart Heuristic: Determine Debit/Credit
                if amt < 0:
                    # Negative values in single column are usually Credits
                    credit_amt = abs(amt)
                elif side == 'dr': 
                    debit_amt = abs(amt)
                elif side == 'cr': 
                    credit_amt = abs(amt)
                else:
                    # Check for Tally custom number format
                    num_fmt = (cell.number_format or '').lower()
                    if 'cr' in num_fmt: 
                        credit_amt = abs(amt)
                    else:
                        # Fallback: Use Keyword Heuristic for Account Names
                        credit_keywords = ['payable', 'capital', 'loan', 'provision', 'reserve', 'equity', 'revenue', 'income', 'tax', 'credit', 'accumulated']
                        if any(k in name.lower() for k in credit_keywords):
                            credit_amt = abs(amt)
                        else:
                            debit_amt = abs(amt)
            elif has_dr_cr:
                dr_val = ws.cell(row=row_idx, column=col_map['debit']+1).value
                cr_val = ws.cell(row=row_idx, column=col_map['credit']+1).value
                debit_amt, side_dummy = self._parse_tally_amount(dr_val)
                credit_amt, side_dummy = self._parse_tally_amount(cr_val)
            
            if debit_amt == 0 and credit_amt == 0: continue
            valid_rows.append({'name': name, 'debit': debit_amt, 'credit': credit_amt})
            all_accounts.append(name)

        # Mapping Check
        mapping_model = self.env['tally.account.mapping']
        for name in all_accounts:
            if not mapping_model.search([('tally_ledger_name', '=', name), ('company_id', '=', self.company_id.id)], limit=1):
                acc_fields = self.env['account.account']._fields
                acc_domain = [('name', '=', name)]
                if 'company_id' in acc_fields: acc_domain.append(('company_id', '=', self.company_id.id))
                elif 'company_ids' in acc_fields: acc_domain.append(('company_ids', 'in', [self.company_id.id]))
                
                odoo_acc = self.env['account.account'].search(acc_domain, limit=1)
                mapping_model.create({'tally_ledger_name': name, 'odoo_account_id': odoo_acc.id if odoo_acc else False, 'company_id': self.company_id.id})

        unmapped = mapping_model.search([('tally_ledger_name', 'in', all_accounts), ('odoo_account_id', '=', False), ('company_id', '=', self.company_id.id)])
        if unmapped:
            return {'name': _('Complete Mappings'), 'type': 'ir.actions.act_window', 'res_model': 'tally.account.mapping', 'view_mode': 'list', 'domain': [('id', 'in', unmapped.ids)], 'target': 'current'}

        ledger_map = {m.tally_ledger_name: m.odoo_account_id.id for m in mapping_model.search([('tally_ledger_name', 'in', all_accounts), ('company_id', '=', self.company_id.id)])}
        batch = self.batch_id or self.env['tally.import.batch'].create({'file_name': self.file_name or 'Tally TB', 'company_id': self.company_id.id})
        
        line_ids = []
        total_debit, total_credit = 0.0, 0.0
        prefix = "Opening" if target_import == 'opening' else "Closing"
        
        for row in valid_rows:
            line_ids.append((0, 0, {'account_id': ledger_map[row['name']], 'name': f"{prefix} - {row['name']}", 'debit': row['debit'], 'credit': row['credit']}))
            total_debit += row['debit']
            total_credit += row['credit']
        
        # Balancing
        currency = self.company_id.currency_id
        diff = currency.round(total_debit - total_credit)
        if diff != 0 and self.auto_balance:
            # Safe domain check for company_id vs company_ids
            acc_fields = self.env['account.account']._fields
            comp_domain = [('code', '=', '999999')]
            if 'company_id' in acc_fields:
                comp_domain.append(('company_id', '=', self.company_id.id))
            elif 'company_ids' in acc_fields:
                comp_domain.append(('company_ids', 'in', [self.company_id.id]))
                
            bal_acc = self.balancing_account_id or self.env['account.account'].search(comp_domain, limit=1)
            if not bal_acc:
                create_vals = {'name': 'Tally Balancing Account', 'code': '999999', 'account_type': 'asset_current'}
                if 'company_id' in acc_fields: create_vals['company_id'] = self.company_id.id
                elif 'company_ids' in acc_fields: create_vals['company_ids'] = [(4, self.company_id.id)]
                bal_acc = self.env['account.account'].create(create_vals)
            line_ids.append((0, 0, {
                'account_id': bal_acc.id, 'name': 'Tally Balancing Difference',
                'debit': abs(diff) if diff < 0 else 0.0, 'credit': diff if diff > 0 else 0.0
            }))
        elif diff != 0 and not self.skip_unbalanced:
             raise UserError(_("Trial Balance is unbalanced by %s. Please check 'Auto-Balance' or provide a Balancing Account.") % diff)

        move_context = {'check_move_validity': False} if (self.skip_unbalanced or self.auto_balance) else {}
        move = self.env['account.move'].with_context(**move_context).create({
            'move_type': 'entry', 'journal_id': self.journal_id.id or self.env['account.journal'].search([('type', '=', 'general')], limit=1).id,
            'date': self.date_to or fields.Date.today(), 'ref': f'Tally {target_import.capitalize()} Import', 'line_ids': line_ids, 'tally_batch_id': batch.id, 'is_tally_import': True
        })
        batch.write({'state': 'done'})
        return {'name': _('Import Successful'), 'type': 'ir.actions.act_window', 'res_model': 'tally.import.batch', 'view_mode': 'form', 'res_id': batch.id, 'target': 'current'}


    def _process_xml(self, file_content):
        # Masters import simplified
        try:
            xml_str = file_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                xml_str = file_content.decode('utf-16')
            except UnicodeDecodeError:
                xml_str = file_content.decode('latin-1', errors='replace')
        
        xml_str = xml_str.strip()
        if '<' in xml_str:
            xml_str = xml_str[xml_str.find('<'):]
            
        try:
            root = ET.fromstring(xml_str)
        except ET.ParseError as e:
            raise UserError(_("Failed to parse XML file. Please check if the file is a valid XML. Error: %s") % str(e))
        for msg in root.findall('.//TALLYMESSAGE'):
            ledger = msg.find('LEDGER')
            if ledger is not None:
                name = ledger.get('NAME') or ledger.findtext('NAME')
                if name:
                    acc_fields = self.env['account.account']._fields
                    acc_domain = [('name', '=', name)]
                    if 'company_id' in acc_fields: acc_domain.append(('company_id', '=', self.company_id.id))
                    elif 'company_ids' in acc_fields: acc_domain.append(('company_ids', 'in', [self.company_id.id]))
                    
                    acc = self.env['account.account'].search(acc_domain, limit=1)
                    if not acc:
                        create_vals = {'name': name, 'code': str(re.sub(r'\D', '', name)[:6]) or '100000', 'account_type': 'asset_current'}
                        if 'company_id' in acc_fields: create_vals['company_id'] = self.company_id.id
                        elif 'company_ids' in acc_fields: create_vals['company_ids'] = [(4, self.company_id.id)]
                        self.env['account.account'].create(create_vals)
        return {'type': 'ir.actions.client', 'tag': 'display_notification', 'params': {'title': _("XML Import Complete"), 'type': 'success'}}
